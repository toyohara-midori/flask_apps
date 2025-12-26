# =======================================================
# 必要な import（順番が超重要）
# =======================================================
from . import db
from flask import render_template, request, send_file
from openpyxl import Workbook
from datetime import date, timedelta
import os
import tempfile

from . import cart_result_bp

# 共通ロジック
from common.cucd_logic import get_cucd_master_tuple
from .db import fetch_cart_stay_all, get_week_calendar
from .db import fetch_total_for_date_and_kbn, fetch_cart_stay_period
from .db import get_category_titles

# 既存のフォーマット系
from .format_common import (
    apply_common_format,
    apply_borders,
    apply_header_color,
    apply_font_style,
    apply_sunday_red
)

from openpyxl.styles import Alignment, Border, Side, Font, Color, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# ① レイアウト生成（値なし）
# ============================================================
def build_base_layout(ws, year, shop_master):
    """
    区分共通のレイアウト（値なし）を作成する
    ・週番号（3行）
    ・日付（4行）
    ・曜日（5行・日曜赤）
    ・店舗CD・店舗名（6行～）
    """
    from .db import get_cucd_info_map
    from common.cucd_logic import get_cucd_name

    cucd_info_map = get_cucd_info_map()

    # --- カレンダー情報取得 ---
    days = get_week_calendar(year)

    # --- 列開始位置 ---
    col_start = 7   # G列から日付
    col_idx = col_start
    prev_weekno = None
    week_start_col = col_start

    # --- 3〜5行目：週番号・日付・曜日 ---
    for weekno, date_obj in days:
        # 週番号セル結合
        if weekno != prev_weekno:
            if prev_weekno is not None:
                ws.merge_cells(start_row=3, start_column=week_start_col,
                               end_row=3, end_column=col_idx - 1)
                ws.cell(row=3, column=week_start_col).value = f"第{prev_weekno}週"
                ws.cell(row=3, column=week_start_col).alignment = Alignment(horizontal="center")
            week_start_col = col_idx
            prev_weekno = weekno

        # 日付
        c_date = ws.cell(row=4, column=col_idx)
        c_date.value = date_obj                 # ← 年付き日付を入れる
        c_date.number_format = "mm/dd"          # ← 表示形式で月日だけにする
        c_date.alignment = Alignment(horizontal="center")

        # 曜日
        weekday = ["月", "火", "水", "木", "金", "土", "日"][date_obj.weekday()]
        c = ws.cell(row=5, column=col_idx)
        c.value = weekday
        c.alignment = Alignment(horizontal="center")

        col_idx += 1

    # 最後の週番号セル結合
    ws.merge_cells(start_row=3, start_column=week_start_col,
                   end_row=3, end_column=col_idx - 1)
    ws.cell(row=3, column=week_start_col).value = f"第{prev_weekno}週"
    ws.cell(row=3, column=week_start_col).alignment = Alignment(horizontal="center")

    # --- 列幅設定 ---
    for c in range(col_start, col_idx):
        ws.column_dimensions[get_column_letter(c)].width = 6

    # --- 店舗CD・店舗名の枠だけ作る（値はのちほど fill_values で入れる） ---
    row = 6
    for cucd, name in shop_master:
        if cucd == "B78":    # ★ 追加：B78 は出力しない
            continue
        ws.cell(row=row, column=2).value = cucd
        ws.cell(row=row, column=3).value = name

        info = cucd_info_map.get(cucd, {})

        fs = info.get("floorSpace")
        ws.cell(row=row, column=4).value = round(float(fs), 1) if fs is not None else ""

        scm = info.get("scmCucd", "")
        SCM_NAME_MAP = {
            "003": "守谷",
            "002": "狭山日高",
        }
        center_name = SCM_NAME_MAP.get(scm, "")
        ws.cell(row=row, column=5).value = center_name

        ws.cell(row=row, column=6).value = info.get("vehicle", "")

        row += 1

    # --- 共通フォーマット ---
    apply_common_format(ws, col_start, col_idx - 1)
    apply_borders(ws, col_start, col_idx - 1)
    apply_header_color(ws, col_start, col_idx - 1)
    apply_font_style(ws)
    apply_sunday_red(ws, col_start, col_idx - 1, days)

    # セル固定
    ws.freeze_panes = "G6"

    return days, col_start, col_idx


# ============================================================
# ② シートコピー
# ============================================================
def duplicate_sheet(wb, base_ws, title):
    """区分1レイアウトシートをコピーして新規シートにする"""

    # まずコピー
    new_ws = wb.copy_worksheet(base_ws)
    new_ws.title = title

    # ★ 列幅をコピーする ★
    for col_letter, col_dim in base_ws.column_dimensions.items():
        if col_dim.width is not None:
            new_ws.column_dimensions[col_letter].width = col_dim.width

    # ★ 行の高さもコピー（必要なら）
    for row_idx, row_dim in base_ws.row_dimensions.items():
        if row_dim.height is not None:
            new_ws.row_dimensions[row_idx].height = row_dim.height

    # 再設定（freeze_panes はコピーされないので必要）
    new_ws.freeze_panes = "G6"

    return new_ws


# ============================================================
# ③ 値埋め込み（cat1〜4、合計）
# ============================================================
def fill_values(ws, days, shop_master, data_dict, kbn_no):
    """
    kbn_no:
      1 → cat1
      2 → cat2
      3 → cat3
      4 → cat4
      "total" → cat1+cat2+cat3+cat4
    """
    col_start = 7
    row_idx = 6

    for cucd, name in shop_master:
        if cucd == "B78":    # ★B78のデータは飛ばす
            continue

        for i, (weekno, date_obj) in enumerate(days):

            col = col_start + i
            ymd = date_obj.strftime("%Y-%m-%d")
            key = f"{cucd}_{ymd}"

            if key not in data_dict:
                value = ""
            else:
                rec = data_dict[key]

                if kbn_no == "total":
                    value = (
                        (rec.get("cat1") or 0) +
                        (rec.get("cat2") or 0) +
                        (rec.get("cat3") or 0) +
                        (rec.get("cat4") or 0)
                    )
                else:
                    raw = rec.get(f"cat{kbn_no}")
                    value = raw if raw is not None else ""

            ws.cell(row=row_idx, column=col).value = value

        row_idx += 1

# ============================================================
# ④ 合計行、前週比行の設定
# ============================================================
def append_summary_rows(ws, days, col_start, col_end, year, kbn_no):
    """
    最下行の下に「合計値」「前週比」を追加する（前週が無い場合は SQL により取得）
    """
    max_row = ws.max_row
    sum_row = max_row + 1
    diff_row = max_row + 2

    # --- ラベル（B空欄、Cに名前） ---
    ws.cell(row=sum_row, column=3).value = "合計値"
    ws.cell(row=diff_row, column=3).value = "前週比"

    ws.cell(row=sum_row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=diff_row, column=3).alignment = Alignment(horizontal="center")

    # --------------------
    # （1）合計値行
    # --------------------
    for i, (weekno, date_obj) in enumerate(days):
        col = col_start + i
        total = 0

        for r in range(6, max_row + 1):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (int, float)):
                total += v

        ws.cell(row=sum_row, column=col).value = total
        ws.cell(row=sum_row, column=col).alignment = Alignment(horizontal="right")

    # -------------------- 
    # （2）前週比行
    # -------------------- 
    for i, (weekno, date_obj) in enumerate(days):
        col = col_start + i

        cur_total = ws.cell(row=sum_row, column=col).value or 0

        # 前週の列が存在する場合（同じ年度内で 7 列前にある）
        prev_col = col - 7
        if prev_col >= col_start:
            prev_total = ws.cell(row=sum_row, column=prev_col).value or 0
            diff = cur_total - prev_total

            cell = ws.cell(row=diff_row, column=col)
            cell.value = diff

            if diff < 0:
                cell.font = Font(color="FF0000")

            cell.alignment = Alignment(horizontal="right")
            continue

        # --------------------------
        # 前週が無い場合（第1週）
        # → 日付-7日 を CartStayCount から直接取得
        # --------------------------
        prev_date = date_obj - timedelta(days=7)
        prev_total = fetch_total_for_date_and_kbn(prev_date, kbn_no)

        diff = cur_total - prev_total

        cell = ws.cell(row=diff_row, column=col)
        cell.value = diff

        if diff < 0:
            cell.font = Font(color="FF0000")

        cell.alignment = Alignment(horizontal="right")


    # ============================================================
    # (3) 見た目整形（罫線・色）
    # ============================================================

    # カラー
    fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # 最終列
    last_col = col_end

    # ------------------------------------------
    # （A）合計行 ＆ 差分行 に色付け（B～最終列）
    # ------------------------------------------
    for col in range(2, last_col + 1):
        ws.cell(row=sum_row,  column=col).fill = fill
        ws.cell(row=diff_row, column=col).fill = fill

    # ------------------------------------------
    # （B）合計行の上を「二重罫線」にする
    # ------------------------------------------
    double = Side(style="double", color="000000")

    for col in range(2, last_col + 1):
        cell = ws.cell(row=sum_row, column=col)
        cell.border = Border(
            top=double,
            left=cell.border.left,
            right=cell.border.right,
            bottom=cell.border.bottom
        )

    # ------------------------------------------
    # （C）太枠を差分行まで伸ばす
    # ------------------------------------------
    thick = Side(style="thick", color="000000")

    # 左枠(B列)
    for row in range(3, diff_row + 1):
        cell = ws.cell(row=row, column=2)
        cell.border = Border(
            left=thick,
            top=cell.border.top,
            bottom=cell.border.bottom,
            right=cell.border.right
        )

    # 右枠（最終列）
    for row in range(3, diff_row + 1):
        cell = ws.cell(row=row, column=last_col)
        cell.border = Border(
            right=thick,
            top=cell.border.top,
            bottom=cell.border.bottom,
            left=cell.border.left
        )

    # 上枠（3行目）→既存維持
    # 下枠（差分行）
    for col in range(2, last_col + 1):
        cell = ws.cell(row=diff_row, column=col)
        cell.border = Border(
            bottom=thick,
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top
        )

    # ============================================================
    # (4) 合計行・差分行 専用罫線
    # ============================================================
    thin = Side(style="dotted", color="000000")
    medium = Side(style="medium", color="000000")

    # ------------------------------------------
    # （1）合計行（sum_row）と前週比行（diff_row）の間に 点線の横罫線
    # ------------------------------------------
    for col in range(2, last_col + 1):
        cell = ws.cell(row=diff_row - 1, column=col)
        cell.border = Border(
            bottom=thin,
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top
        )

    # ------------------------------------------
    # （2）合計行・前週比行の全曜日に 縦の点線（D列～最終列）
    # ------------------------------------------
    for col in range(col_start, last_col + 1):
        # 合計行
        cell = ws.cell(row=sum_row, column=col)
        cell.border = Border(
            left=thin,
            right=thin,
            top=cell.border.top,
            bottom=cell.border.bottom
        )

        # 前週比行
        cell = ws.cell(row=diff_row, column=col)
        cell.border = Border(
            left=thin,
            right=thin,
            top=cell.border.top,
            bottom=cell.border.bottom
        )

    # ------------------------------------------
    # （3）日曜 → 月曜 の境界に medium 線（縦）
    # days: [(weekno, date), ...] の配列
    # 曜日 index: 月=0, 火=1 ... 日=6
    # ------------------------------------------
    for i, (weekno, date_obj) in enumerate(days):
        col = col_start + i
        # 月曜列の場合のみ「その直前」が日曜
        if date_obj.weekday() == 0 and col > col_start:
            # 合計行
            cell = ws.cell(row=sum_row, column=col)
            cell.border = Border(
                left=medium,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            # 前週比行
            cell = ws.cell(row=diff_row, column=col)
            cell.border = Border(
                left=medium,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )

    # ============================================================
    # (4) 最終：合計行の上に double 線を “再適用” （上書き保護）
    #           合計行・差分行の最終列の右側に太線を強制適用
    # ============================================================

    double = Side(style="double", color="000000")

    for col in range(2, last_col + 1):
        cell = ws.cell(row=sum_row, column=col)

        # 既存の左右・下線は保持しつつ、上だけ double にする
        cell.border = Border(
            top=double,
            left=cell.border.left,
            right=cell.border.right,
            bottom=cell.border.bottom
        )

    thick = Side(style="thick", color="000000")

    for row in (sum_row, diff_row):
        cell = ws.cell(row=row, column=last_col)
        cell.border = Border(
            right=thick,
            top=cell.border.top,
            bottom=cell.border.bottom,
            left=cell.border.left,
        )

# ============================================================
# 共通 Excel 生成関数
# ============================================================
def build_excel_workbook(year, shop_master, data_dict):
    wb = Workbook()
    base_ws = wb.active
    base_ws.title = "区分1"

    # 区分名
    title_map = get_category_titles()

    # レイアウト作成
    days, col_start, col_end = build_base_layout(base_ws, year, shop_master)

    # タイトル
    base_ws["B2"] = title_map[1]
    base_ws["B2"].font = Font(name="Meiryo UI", size=14, bold=True)

    unit_cell = base_ws.cell(row=2, column=col_end - 1)
    unit_cell.value = "単位：台"
    unit_cell.font = Font(name="Meiryo UI", size=11)
    unit_cell.alignment = Alignment(horizontal="center")

    # コピーして各シート作成
    ws2 = duplicate_sheet(wb, base_ws, "区分2")
    ws3 = duplicate_sheet(wb, base_ws, "区分3")
    ws4 = duplicate_sheet(wb, base_ws, "区分4")
    ws_total = duplicate_sheet(wb, base_ws, "滞留カゴ車台数実績表")

    ws2["B2"] = title_map[2]
    ws3["B2"] = title_map[3]
    ws4["B2"] = title_map[4]
    ws_total["B2"] = "滞留カゴ車台数実績表"

    # 値の埋め込み
    fill_values(base_ws, days, shop_master, data_dict, 1)
    append_summary_rows(base_ws, days, col_start, col_end - 1, year, 1)

    fill_values(ws2, days, shop_master, data_dict, 2)
    append_summary_rows(ws2, days, col_start, col_end - 1, year, 2)

    fill_values(ws3, days, shop_master, data_dict, 3)
    append_summary_rows(ws3, days, col_start, col_end - 1, year, 3)

    fill_values(ws4, days, shop_master, data_dict, 4)
    append_summary_rows(ws4, days, col_start, col_end - 1, year, 4)

    fill_values(ws_total, days, shop_master, data_dict, "total")
    append_summary_rows(ws_total, days, col_start, col_end - 1, year, "total")

    return wb

# ============================================================
# ④ ルート
# ============================================================
@cart_result_bp.route("/")
def index():
    return render_template("cart_result.html")


# ============================================================
# ⑤ Excel出力（高速版）
# ============================================================
@cart_result_bp.route("/export_excel", methods=["POST"])
def export_excel():
    year_type = request.form.get("year_type")
    today = date.today()
    year = today.year if year_type == "current" else today.year - 1

    title_map = get_category_titles()

    # 店舗マスター（タプル形式）
    shop_master = get_cucd_master_tuple()

    # 区分データ1回取得
    data_dict = fetch_cart_stay_all(year)

    # Excelブック作成
    wb = Workbook()
    base_ws = wb.active
    base_ws.title = "区分1"

    # ---- 区分1レイアウト作成 ----
    days, col_start, col_end = build_base_layout(base_ws, year, shop_master)
    
    # 2行目B列タイトル
    base_ws["B2"] = title_map[1]
    base_ws["B2"].font = Font(name="Meiryo UI", size=14, bold=True)
    # 2行目最終列　単位
    unit_cell = base_ws.cell(row=2, column=col_end - 1)
    unit_cell.value = "単位：台"
    unit_cell.font = Font(name="Meiryo UI", size=11)
    unit_cell.alignment = Alignment(horizontal="center")

    # ---- 区分2〜4、合計シートをコピーして作成 ----
    ws2 = duplicate_sheet(wb, base_ws, "区分2")
    ws3 = duplicate_sheet(wb, base_ws, "区分3")
    ws4 = duplicate_sheet(wb, base_ws, "区分4")
    ws_total = duplicate_sheet(wb, base_ws, "滞留カゴ車台数実績表")

    # ---- 各シートのB2セルに、シートタイトルを設定----
    ws2["B2"] = title_map[2]
    ws3["B2"] = title_map[3]
    ws4["B2"] = title_map[4]
    ws_total["B2"] = "滞留カゴ車台数実績表"

    # ---- 値を埋め込む（高速）----
    fill_values(base_ws, days, shop_master, data_dict, 1)
    append_summary_rows(base_ws, days, col_start, col_end - 1, year, 1)
    fill_values(ws2,    days, shop_master, data_dict, 2)
    append_summary_rows(ws2,  days, col_start, col_end - 1, year, 2)
    fill_values(ws3,    days, shop_master, data_dict, 3)
    append_summary_rows(ws3,  days, col_start, col_end - 1, year, 3)
    fill_values(ws4,    days, shop_master, data_dict, 4)
    append_summary_rows(ws4,  days, col_start, col_end - 1, year, 4)
    fill_values(ws_total, days, shop_master, data_dict, "total")
    append_summary_rows(ws_total, days, col_start, col_end - 1, year, "total")

    # ---- ファイル名・保存 ----
    filename = f"{year}年度_滞留カゴ車台数実績表.xlsx"
    temp_path = os.path.join(tempfile.gettempdir(), filename)
    wb.save(temp_path)

    return send_file(temp_path, as_attachment=True)

# ============================================================
# Excel出力（zipで今年度と2週間分の2ファイル出力版）
# ============================================================
@cart_result_bp.route("/export_excel_zip", methods=["POST"])
def export_excel_zip():
    today = date.today()
    ymd = today.strftime("%Y%m%d")
    year = today.year  # 今年度出力

    # 店舗マスター
    shop_master = get_cucd_master_tuple()

    # -------------------------
    # ① 年度版 Excel を生成
    # -------------------------
    data_full = fetch_cart_stay_all(year)
    wb_full = build_excel_workbook(year, shop_master, data_full)
    buf_full = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb_full.save(buf_full.name)

    # -------------------------
    # ② 2週間版 Excel を生成
    # -------------------------
    start_2w = today - timedelta(days=13)
    end_2w = today

    data_2w = fetch_cart_stay_period(start_2w, end_2w)
    wb_2w = create_excel_two_weeks()
    fname_2w = f"2週間滞留カゴ車台数実績表({ymd}).xlsx"
    buf_2w = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb_2w.save(buf_2w.name)

    # -------------------------
    # ③ ZIP 作成
    # -------------------------
    zip_path = tempfile.NamedTemporaryFile(delete=False, suffix=".zip").name
    import zipfile
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(buf_full.name, arcname=f"{year}年度_滞留カゴ車台数実績表.xlsx")
        z.write(buf_2w.name, arcname=fname_2w)

    return send_file(zip_path,
                     as_attachment=True,
                     download_name=f"滞留カゴ車集計_{year}年度版＋2週間版_{ymd}.zip")

# ============================================================
# Excel出力（2週間ファイル）
# ============================================================
def build_base_layout_period(ws, start_date, end_date, shop_master):
    """
    指定期間（start_date〜end_date）のみのレイアウトを作る。
    ・週番号（必要な分だけ）
    ・日付（期間の分だけ）
    ・曜日
    ・店舗CD・店舗名
    """
    from common.db_connection import get_connection
    from .db import get_cucd_info_map
    from common.cucd_logic import get_cucd_name

    cucd_info_map = get_cucd_info_map()

    # --------------------------
    # ① 期間内の日付リスト作成
    # --------------------------
    cur_date = start_date
    days = []

    with get_connection("SQLS08-14") as conn:
        cur = conn.cursor()

        while cur_date <= end_date:
            cur.execute("""
                SELECT weekno 
                FROM dba.weekno2 
                WHERE ? BETWEEN date_s AND date_e
            """, (cur_date,))
            row = cur.fetchone()
            weekno = row[0] if row else 0   # 取れないことは基本ない
            
            days.append((weekno, cur_date))
            cur_date += timedelta(days=1)

    # --------------------------
    # ② カレンダー描画（年度版と同じ方式）
    # --------------------------
    col_start = 7   # G列から日付
    col_idx = col_start
    prev_weekno = None
    week_start_col = col_start

    for (weekno, date_obj) in days:
        # 週番号セル結合
        if weekno != prev_weekno:
            if prev_weekno is not None:
                ws.merge_cells(start_row=3, start_column=week_start_col,
                               end_row=3, end_column=col_idx - 1)
                ws.cell(row=3, column=week_start_col).value = f"第{prev_weekno}週"
                ws.cell(row=3, column=week_start_col).alignment = Alignment(horizontal="center")
            week_start_col = col_idx
            prev_weekno = weekno

        # 日付
        c_date = ws.cell(row=4, column=col_idx)
        c_date.value = date_obj
        c_date.number_format = "mm/dd"
        c_date.alignment = Alignment(horizontal="center")

        # 曜日
        weekday = ["月", "火", "水", "木", "金", "土", "日"][date_obj.weekday()]
        c = ws.cell(row=5, column=col_idx)
        c.value = weekday
        c.alignment = Alignment(horizontal="center")

        col_idx += 1

    # 最後の週番号セル結合
    ws.merge_cells(start_row=3, start_column=week_start_col,
                   end_row=3, end_column=col_idx - 1)
    ws.cell(row=3, column=week_start_col).value = f"第{prev_weekno}週"
    ws.cell(row=3, column=week_start_col).alignment = Alignment(horizontal="center")

    # 列幅
    for c in range(col_start, col_idx):
        ws.column_dimensions[get_column_letter(c)].width = 6

    # 店舗一覧
    row = 6
    for cucd, name in shop_master:
        if cucd == "B78":
            continue
        ws.cell(row=row, column=2).value = cucd
        ws.cell(row=row, column=3).value = name

        info = cucd_info_map.get(cucd, {})

        fs = info.get("floorSpace")
        ws.cell(row=row, column=4).value = round(float(fs), 1) if fs is not None else ""

        scm = info.get("scmCucd", "")
        SCM_NAME_MAP = {
            "003": "守谷",
            "002": "狭山日高",
        }
        center_name = SCM_NAME_MAP.get(scm, "")
        ws.cell(row=row, column=5).value = center_name

        ws.cell(row=row, column=6).value = info.get("vehicle", "")

        row += 1

    # 共通フォーマット適用
    apply_common_format(ws, col_start, col_idx - 1)
    apply_borders(ws, col_start, col_idx - 1)
    apply_header_color(ws, col_start, col_idx - 1)
    apply_font_style(ws)
    apply_sunday_red(ws, col_start, col_idx - 1, days)

    # セル固定
    ws.freeze_panes = "G6"

    return days, col_start, col_idx

def create_excel_two_weeks():
    today = date.today()
    start_2w = today - timedelta(days=13)
    end_2w   = today

    # 店舗マスター
    shop_master = get_cucd_master_tuple()

    # 2週間分のデータ
    data_dict = fetch_cart_stay_period(start_2w, end_2w)

    # Excelブック生成
    wb = Workbook()
    ws_base = wb.active
    ws_base.title = "区分1"

    # ★ 2週間版のレイアウト（列が14列だけ）
    days, col_start, col_end = build_base_layout_period(ws_base, start_2w, end_2w, shop_master)

    # 区分タイトル
    title_map = get_category_titles()

    ws_base["B2"] = title_map[1]

    # シート複製（区分2〜4 & 合計）
    ws2 = duplicate_sheet(wb, ws_base, "区分2")
    ws3 = duplicate_sheet(wb, ws_base, "区分3")
    ws4 = duplicate_sheet(wb, ws_base, "区分4")
    ws_total = duplicate_sheet(wb, ws_base, "滞留カゴ車台数実績表")

    ws2["B2"] = title_map[2]
    ws3["B2"] = title_map[3]
    ws4["B2"] = title_map[4]
    ws_total["B2"] = "滞留カゴ車台数実績表"

    # 値埋め込み
    fill_values(ws_base, days, shop_master, data_dict, 1)
    append_summary_rows(ws_base, days, col_start, col_end - 1, None, 1)

    fill_values(ws2, days, shop_master, data_dict, 2)
    append_summary_rows(ws2, days, col_start, col_end - 1, None, 2)

    fill_values(ws3, days, shop_master, data_dict, 3)
    append_summary_rows(ws3, days, col_start, col_end - 1, None, 3)

    fill_values(ws4, days, shop_master, data_dict, 4)
    append_summary_rows(ws4, days, col_start, col_end - 1, None, 4)

    fill_values(ws_total, days, shop_master, data_dict, "total")
    append_summary_rows(ws_total, days, col_start, col_end - 1, None, "total")

    return wb

# ===============================
# データ照会画面（GET）
# ===============================
@cart_result_bp.route("/cart_result_disp")
def cart_result_disp():
    from datetime import date, timedelta

    today = date.today()
    # 今日から4週間前
    start_date = today - timedelta(weeks=4)
    end_date = today

    # yyyy-MM-dd の文字列にしてテンプレートへ渡す
    ctx = {
        "default_start": start_date.strftime("%Y-%m-%d"),
        "default_end": end_date.strftime("%Y-%m-%d"),
    }

    return render_template("cart_result_disp.html", **ctx)

# ============================================
#  データ取得 API（表示ボタン）
# ============================================
@cart_result_bp.route("/get_data", methods=["POST"])
def get_data():
    from flask import request, jsonify
    from datetime import datetime
    from common.db_connection import get_connection

    start_date = request.form.get("start_date")
    end_date   = request.form.get("end_date")
    disp_type  = request.form.get("disp_type")   # total / 1 / 2 / 3 / 4

    # 文字列 → date
    try:
        sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        ed = datetime.strptime(end_date, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "日付の形式が不正です"}), 400

    # ---- 1) CartStayCount からデータ取得 ----
    with get_connection("SQLS08-14") as conn:
        cur = conn.cursor()

        # 同一CUCD/日付のレコードが複数存在することがあるので、最新のレコードのみ取得
        sql = """
            SELECT
                c.cucd,
                c.idleDate,
                c.cat1,
                c.cat2,
                c.cat3,
                c.cat4
            FROM DBA.CartStayCount c
            INNER JOIN (
                SELECT
                    cucd,
                    idleDate,
                    MAX(rgdt) AS max_rgdt
                FROM DBA.CartStayCount
                WHERE idleDate BETWEEN ? AND ?
                GROUP BY cucd, idleDate
            ) m
                ON  m.cucd     = c.cucd
                AND m.idleDate = c.idleDate
                AND m.max_rgdt = c.rgdt
            WHERE c.idleDate BETWEEN ? AND ?
        """
        cur.execute(sql, (sd, ed, sd, ed))
        rows = cur.fetchall()

    # ---- 2) データを辞書化（(cucd, date) → 値） ----
    data = []
    for r in rows:
        cucd = str(r.cucd).strip()
        d = r.idleDate
        if hasattr(d, "date"):
            d = d.date()

        rec = {
            "cucd": cucd,
            "date": d.strftime("%Y-%m-%d"),
            "cat1": r.cat1 or 0,
            "cat2": r.cat2 or 0,
            "cat3": r.cat3 or 0,
            "cat4": r.cat4 or 0,
        }

        # ---- 区分ごとの値の抽出 ----
        if disp_type == "total":
            rec["value"] = (r.cat1 or 0) + (r.cat2 or 0) + (r.cat3 or 0) + (r.cat4 or 0)
        else:
            k = f"cat{disp_type}"
            rec["value"] = rec[k]

        data.append(rec)

    return jsonify({"result": data})

#--- 実績表表示画面で、店舗一覧を作るためのAPI ---
@cart_result_bp.route("/get_shop_master")
def get_shop_master():
    from common.cucd_logic import get_cucd_master_tuple
    shops = get_cucd_master_tuple()   # [(cucd, name), ...]
    return {"shops": shops}
