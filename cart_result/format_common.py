from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter


# -----------------------------------------
# 共通フォーマット
# -----------------------------------------
def apply_common_format(ws, col_start, col_end):
    """セル結合・列幅・固定位置"""

    # 列幅
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 22
    for c in range(col_start, col_end + 1):
        ws.column_dimensions[get_column_letter(c)].width = 7

    # B3〜B5
    ws.merge_cells("B3:B5")
    ws["B3"].value = "店番"
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

    # C3〜C5
    ws.merge_cells("C3:C5")
    ws["C3"].value = "店舗名"
    ws["C3"].alignment = Alignment(horizontal="center", vertical="center")

    # 2行目の高さを設定
    ws.row_dimensions[2].height = 19.5

    # ★ セル固定
    #    (でもこれはコピー後のシートには適用されないので、コピー後に再設定が必要)
    ws.freeze_panes = "D6"

# -----------------------------------------
# 罫線設定
# -----------------------------------------
def apply_borders(ws, col_start, col_end):
    """太枠・縦点線・横点線・週境界（日→月）"""

    thin = Side(style="dotted", color="000000")
    thick = Side(style="thick", color="000000")
    medium = Side(style="medium", color="000000")

    max_row = ws.max_row
    sum_row = ws.max_row - 1   # 合計行
    diff_row = ws.max_row      # 前週比行

    # ① 外枠（上・下・左・右 全て太線）

    # 上枠
    for col in range(2, col_end + 1):
        cell = ws.cell(row=3, column=col)
        cell.border = Border(
            top=thick,
            left=cell.border.left,
            right=cell.border.right,
            bottom=cell.border.bottom,
        )

    # 下枠
    for col in range(2, col_end + 1):
        # ★ 合計行・差分行には bottom=thick を適用しない
        if diff_row == ws.max_row and (ws.max_row == diff_row or ws.max_row == sum_row):
            continue
        cell = ws.cell(row=max_row, column=col)
        cell.border = Border(
            bottom=thick,
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
        )

    # 左枠（B列）
    for row in range(3, max_row + 1):
        cell = ws.cell(row=row, column=2)
        cell.border = Border(
            left=thick,
            top=cell.border.top,
            bottom=cell.border.bottom,
            right=cell.border.right,
        )

    # ★右枠（最終列）
    for row in range(3, max_row + 1):
        cell = ws.cell(row=row, column=col_end)
        cell.border = Border(
            right=thick,
            top=cell.border.top,
            bottom=cell.border.bottom,
            left=cell.border.left,
        )

    # ② 縦点線（4行目〜最終行）
    for col in range(col_start, col_end + 1):
        for row in range(4, max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left=thin,
                right=thin,
                top=cell.border.top,
                bottom=cell.border.bottom
            )

    # ③ 横点線（4〜5行間）
    for col in range(2, col_end + 1):
        cell = ws.cell(row=4, column=col)
        cell.border = Border(
            bottom=thin,
            top=cell.border.top,
            left=cell.border.left,
            right=cell.border.right,
        )

    # ④ 横点線（6行目〜最終行）
    for row in range(6, max_row + 1):
        for col in range(2, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                top=thin,
                left=cell.border.left,
                right=cell.border.right,
                bottom=cell.border.bottom,
            )

    # ⑤ 日曜→月曜の境界に medium 線
    # ⑤ 日曜→月曜の境界に太線（曜日ベース）
    # days は build_base_layout / build_base_layout_period の戻り値を利用
    # → apply_borders の引数に days を追加していないので、
    #    ここで date_obj を取得する必要がある

    # ★ days は ws 上の 4行目（mm/dd）の値から取得できる
    #   （レイアウト生成時に書き込んだ日付セルを参照する）
    for col in range(col_start, col_end + 1):
        date_cell = ws.cell(row=4, column=col).value

        # 日付セルが date 型であることを確認
        if hasattr(date_cell, "weekday") and date_cell.weekday() == 0:
            # ここが月曜日の列 → 左側に太線
            for row in range(4, max_row + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=thick,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )

    # ★ 最終：右枠の太線をもう一度上書きして復活させる
    for row in range(3, max_row + 1):
        cell = ws.cell(row=row, column=col_end)
        cell.border = Border(
            right=thick,
            top=cell.border.top,
            bottom=cell.border.bottom,
            left=cell.border.left,
        )

# -----------------------------------------
# 見出しのセル色
# -----------------------------------------
def apply_header_color(ws, col_start, col_end):
    fill = PatternFill("solid", fgColor="C6D9F1")

    for row in [3, 4, 5]:
        for col in range(2, col_end + 1):
            ws.cell(row=row, column=col).fill = fill


# -----------------------------------------
# フォント設定
# -----------------------------------------
def apply_font_style(ws):
    # B2（タイトル）
    title_cell = ws["B2"]
    title_cell.font = Font(name="Meiryo UI", size=14, bold=True)

    # その他
    for row in ws.iter_rows():
        for cell in row:
            if cell is not title_cell:
                cell.font = Font(name="Meiryo UI", size=11)


# -----------------------------------------
# 日曜日を赤字に
# -----------------------------------------
def apply_sunday_red(ws, col_start, col_end, days):
    for i, (weekno, date_obj) in enumerate(days):
        if date_obj.weekday() == 6:  # 日曜
            col = col_start + i
            ws.cell(5, col).font = Font(color="FF0000")
